from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status as drf_status
from django.utils import timezone
from django.db.models import Sum, Count, F, Q, Case, When, Value, DecimalField
from django.db.models.functions import TruncDate

from apps.users.permissions import CanViewOnly
from apps.sales.models import Sale, SaleItem, Return
from apps.orders.models import Order, PurchaseOrder
from apps.inventory.models import ProductVariant, StockMovement, Product
from apps.cash.models import CashSession

import io

# Etiquetas de métodos de pago base (el campo Sale.payment_method no usa choices,
# así que get_payment_method_display() no existe — usamos este dict como fallback).
_BASE_PAYMENT_LABELS = {
    "cash":      "Efectivo",
    "transfer":  "Transferencia",
    "debit":     "Débito",
    "credit":    "Crédito",
    "nequi":     "Nequi",
    "daviplata": "Daviplata",
    "other":     "Otro",
}


def _payment_label(key: str) -> str:
    """Devuelve la etiqueta legible de un método de pago dado su clave."""
    return _BASE_PAYMENT_LABELS.get(key, key)


def _pct_change(current, previous):
    """Calcula el % de cambio entre dos valores. Retorna None si el anterior es 0."""
    if not previous:
        return None
    return round(((current - previous) / previous) * 100, 1)


class DashboardView(APIView):
    """Dashboard con KPIs, gráfica de ventas y alertas de stock."""
    permission_classes = [CanViewOnly]

    def get(self, request):
        from datetime import timedelta
        try:
            return self._build_response(timedelta)
        except Exception as exc:
            return Response(
                {"detail": "Error al cargar el dashboard.", "error": str(exc)},
                status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def _build_response(self, timedelta):
        today = timezone.localdate()
        yesterday = today - timedelta(days=1)
        thirty_days_start = today - timedelta(days=29)   # date — ventana uniforme de 30 días
        thirty_days_ago   = timezone.now() - timedelta(days=30)  # datetime — para restock alerts

        # ── Ventas hoy y ayer (2 queries, cada una con aggregate combinado) ──
        today_agg = Sale.objects.filter(created_at__date=today).aggregate(
            t=Sum("total"), cnt=Count("id")
        )
        today_revenue = float(today_agg["t"] or 0)
        today_count   = today_agg["cnt"] or 0

        yesterday_agg = Sale.objects.filter(created_at__date=yesterday).aggregate(
            t=Sum("total"), cnt=Count("id")
        )
        yesterday_revenue = float(yesterday_agg["t"] or 0)
        yesterday_count   = yesterday_agg["cnt"] or 0

        # ── Pedidos nuevos ───────────────────────────────────────────────────
        new_orders           = Order.objects.filter(status="new").count()
        yesterday_new_orders = Order.objects.filter(
            status="new", created_at__date=yesterday
        ).count()

        # ── Stock bajo / agotado ─────────────────────────────────────────────
        low_stock_count    = ProductVariant.objects.filter(
            is_active=True, stock__gt=0, stock__lte=F("product__min_stock")
        ).count()
        out_of_stock_count = ProductVariant.objects.filter(is_active=True, stock=0).count()

        # ── Caja actual (1 query con Case/When en lugar de 3 separadas) ──────
        session   = CashSession.objects.filter(date=today, status="open").first()
        cash_data = None
        if session:
            agg = session.movements.aggregate(
                income=Sum(
                    Case(When(type="income", then="amount"),
                         default=Value(0), output_field=DecimalField())
                ),
                expense=Sum(
                    Case(When(type="expense", then="amount"),
                         default=Value(0), output_field=DecimalField())
                ),
                refund=Sum(
                    Case(When(type="refund", then="amount"),
                         default=Value(0), output_field=DecimalField())
                ),
            )
            income  = float(agg["income"]  or 0)
            expense = float(agg["expense"] or 0)
            refund  = float(agg["refund"]  or 0)
            cash_data = {
                "session_id":     session.id,
                "opening_amount": str(session.opening_amount),
                "current_cash":   float(session.opening_amount) + income - expense - refund,
                "total_income":   income,
                "total_expense":  expense,
                "total_refund":   refund,
            }

        # ── Gráfica de ventas — últimos 30 días (1 query) ────────────────────
        tz = timezone.get_current_timezone()
        sales_day_map = {
            str(row["day"]): float(row["t"])
            for row in (
                Sale.objects
                .filter(created_at__date__gte=thirty_days_start)
                .annotate(day=TruncDate("created_at", tzinfo=tz))
                .values("day")
                .annotate(t=Sum("total"))
            )
        }
        sales_chart = [
            {"date": str(today - timedelta(days=i)),
             "total": sales_day_map.get(str(today - timedelta(days=i)), 0.0)}
            for i in range(29, -1, -1)
        ]

        # ── Gráfica de pedidos — últimos 14 días (1 query) ───────────────────
        fourteen_days_start = today - timedelta(days=13)
        orders_day_map = {
            str(row["day"]): row["cnt"]
            for row in (
                Order.objects
                .filter(created_at__date__gte=fourteen_days_start)
                .annotate(day=TruncDate("created_at", tzinfo=tz))
                .values("day")
                .annotate(cnt=Count("id"))
            )
        }
        orders_chart = [
            {"date": str(today - timedelta(days=i)),
             "total": orders_day_map.get(str(today - timedelta(days=i)), 0)}
            for i in range(13, -1, -1)
        ]

        # ── Top 5 productos más vendidos — últimos 30 días (1 query) ─────────
        top_products_list = [
            {
                "product_id": p["variant__product__id"],
                "name":       p["variant__product__name"],
                "category":   p["variant__product__category__name"] or "Sin categoría",
                "units_sold": p["units_sold"],
                "revenue":    float(p["revenue"]),
            }
            for p in (
                SaleItem.objects
                .filter(sale__created_at__date__gte=thirty_days_start)
                .values("variant__product__id", "variant__product__name",
                        "variant__product__category__name")
                .annotate(units_sold=Sum("quantity"), revenue=Sum("subtotal"))
                .order_by("-units_sold")[:5]
            )
        ]

        # ── Alertas de stock bajo (variantes ≤ mínimo, incluye agotadas) ─────
        stock_alerts = [
            {
                "variant_id":    v.id,
                "product_id":    v.product.id,
                "product_name":  v.product.name,
                "size":          v.size,
                "color":         v.color,
                "current_stock": v.stock,
                "min_stock":     v.product.min_stock,
                "urgency":       "out" if v.stock == 0 else "critical" if v.stock <= 1 else "low",
            }
            for v in (
                ProductVariant.objects
                .filter(is_active=True, stock__lte=F("product__min_stock"))
                .select_related("product")
                .order_by("stock")[:20]
            )
        ]

        # ── Pedidos recientes (últimos 5) ────────────────────────────────────
        recent_orders = [
            {
                "id":            o.id,
                "number":        o.number,
                "status":        o.status,
                "total":         float(o.total) if o.total else 0,
                "customer_name": o.customer_name or "Cliente anónimo",
            }
            for o in Order.objects.order_by("-created_at")[:5]
        ]

        # ── Predicción de reabastecimiento ───────────────────────────────────
        restock_alerts = _calculate_restock_alerts(thirty_days_ago)

        # Marcar alertas que ya tienen una OC activa (pending o partial)
        # Evita que se creen OC duplicadas para la misma variante.
        alert_variant_ids = [a["variant_id"] for a in restock_alerts]
        active_ocs = {
            oc["variant_id"]: {"number": oc["number"], "id": oc["id"]}
            for oc in PurchaseOrder.objects.filter(
                variant_id__in=alert_variant_ids,
                status__in=[PurchaseOrder.Status.PENDING, PurchaseOrder.Status.PARTIAL],
            ).values("variant_id", "number", "id")
        }
        for alert in restock_alerts:
            oc = active_ocs.get(alert["variant_id"])
            alert["has_pending_oc"]    = oc is not None
            alert["pending_oc_number"] = oc["number"] if oc else None
            alert["pending_oc_id"]     = oc["id"]     if oc else None

        # ── Compras pendientes (badge global) ────────────────────────────────
        pending_purchases = PurchaseOrder.objects.filter(
            status__in=[PurchaseOrder.Status.PENDING, PurchaseOrder.Status.PARTIAL]
        ).count()

        return Response({
            "today_revenue":         today_revenue,
            "today_sales_count":     today_count,
            "yesterday_revenue":     yesterday_revenue,
            "yesterday_sales_count": yesterday_count,
            "revenue_change":        _pct_change(today_revenue, yesterday_revenue),
            "sales_count_change":    _pct_change(today_count, yesterday_count),
            "new_orders":            new_orders,
            "yesterday_new_orders":  yesterday_new_orders,
            "low_stock_count":       low_stock_count,
            "out_of_stock_count":    out_of_stock_count,
            "cash_session":          cash_data,
            "sales_chart":           sales_chart,
            "orders_chart":          orders_chart,
            "top_products":          top_products_list,
            "stock_alerts":          stock_alerts,
            "recent_orders":         recent_orders,
            "restock_alerts":        restock_alerts,
            "pending_purchases":     pending_purchases,
        })


def _calculate_restock_alerts(thirty_days_ago):
    """
    Calcula variantes con estimado < 10 días de stock.
    Usa una sola query con anotaciones en lugar de N queries (una por variante).
    """
    # Anota cada variante con las unidades vendidas en los últimos 30 días.
    # Las cantidades de venta en StockMovement son negativas (type="sale").
    variants = (
        ProductVariant.objects
        .filter(is_active=True, stock__gt=0)
        .select_related("product")
        .annotate(
            units_sold_30d=Sum(
                Case(
                    When(
                        movements__type="sale",
                        movements__created_at__gte=thirty_days_ago,
                        then="movements__quantity",
                    ),
                    default=Value(0),
                    output_field=DecimalField(),
                )
            )
        )
        .filter(units_sold_30d__lt=0)  # solo las que tuvieron ventas
    )

    alerts = []
    for v in variants:
        daily_avg = abs(float(v.units_sold_30d)) / 30
        days_remaining = v.stock / daily_avg
        if days_remaining < 10:
            alerts.append({
                "variant_id": v.id,
                "product_id": v.product.id,
                "product_name": v.product.name,
                "size": v.size,
                "color": v.color,
                "current_stock": v.stock,
                "days_remaining": round(days_remaining, 1),
            })
    return sorted(alerts, key=lambda x: x["days_remaining"])


class InventoryAlertsView(APIView):
    """Alertas de stock y predicciones de reabastecimiento — endpoint dedicado para el módulo de Inventario."""
    permission_classes = [CanViewOnly]

    def get(self, request):
        from datetime import timedelta
        thirty_days_ago = timezone.now() - timedelta(days=30)

        stock_alerts = [
            {
                "variant_id":    v.id,
                "product_id":    v.product.id,
                "product_name":  v.product.name,
                "size":          v.size,
                "color":         v.color,
                "current_stock": v.stock,
                "min_stock":     v.product.min_stock,
                "urgency":       "out" if v.stock == 0 else "critical" if v.stock <= 1 else "low",
            }
            for v in (
                ProductVariant.objects
                .filter(is_active=True, stock__lte=F("product__min_stock"))
                .select_related("product")
                .order_by("stock")[:50]
            )
        ]

        restock_alerts = _calculate_restock_alerts(thirty_days_ago)

        alert_variant_ids = [a["variant_id"] for a in restock_alerts]
        active_ocs = {}
        if alert_variant_ids:
            active_ocs = {
                oc["variant_id"]: {"number": oc["number"], "id": oc["id"]}
                for oc in PurchaseOrder.objects.filter(
                    variant_id__in=alert_variant_ids,
                    status__in=[PurchaseOrder.Status.PENDING, PurchaseOrder.Status.PARTIAL],
                ).values("variant_id", "number", "id")
            }
        for alert in restock_alerts:
            oc = active_ocs.get(alert["variant_id"])
            alert["has_pending_oc"]    = oc is not None
            alert["pending_oc_number"] = oc["number"] if oc else None
            alert["pending_oc_id"]     = oc["id"]     if oc else None

        return Response({
            "stock_alerts":   stock_alerts,
            "restock_alerts": restock_alerts,
        })


class SalesReportView(APIView):
    """Reporte de ventas con filtros — devuelve todos los datos para los gráficos del frontend."""
    permission_classes = [CanViewOnly]

    def get(self, request):
        from datetime import date, timedelta
        from django.db.models import Q

        from_date_str = request.query_params.get("from_date")
        to_date_str   = request.query_params.get("to_date")
        payment = request.query_params.get("payment_method")

        if from_date_str and to_date_str:
            try:
                start_date = date.fromisoformat(from_date_str)
                end_date   = date.fromisoformat(to_date_str)
                days = max(1, (end_date - start_date).days + 1)
            except ValueError:
                end_date   = timezone.now().date()
                days = int(request.query_params.get("days", 30))
                start_date = end_date - timedelta(days=days - 1)
        else:
            days = int(request.query_params.get("days", 30))
            end_date   = timezone.now().date()
            start_date = end_date - timedelta(days=days - 1)

        qs = Sale.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
        if payment:
            qs = qs.filter(payment_method=payment)

        # ── Totales ────────────────────────────────────────
        totals = qs.aggregate(total=Sum("total"), count=Count("id"))
        gross_revenue = float(totals["total"] or 0)
        total_sales   = totals["count"] or 0

        # Restar devoluciones del período
        returns_agg = Return.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            type=Return.ReturnType.RETURN,
        ).aggregate(
            refunded=Sum(F("returned_price") * F("returned_quantity"), output_field=DecimalField())
        )
        total_refunded = float(returns_agg["refunded"] or 0)
        total_revenue = max(0, gross_revenue - total_refunded)

        daily_avg     = round(total_revenue / days, 2) if days else 0

        # ── Ventas por día (1 query con TruncDate) ─────────
        tz = timezone.get_current_timezone()
        sales_by_day_qs = (
            qs.annotate(day=TruncDate("created_at", tzinfo=tz))
            .values("day")
            .annotate(t=Sum("total"))
        )
        day_map = {str(row["day"]): float(row["t"]) for row in sales_by_day_qs}
        sales_by_day = [
            {"date": str(end_date - timedelta(days=i)), "total": day_map.get(str(end_date - timedelta(days=i)), 0.0)}
            for i in range(days - 1, -1, -1)
        ]

        # ── Top 5 productos ────────────────────────────────
        top_products = (
            SaleItem.objects.filter(sale__in=qs)
            .values("variant__product__id", "variant__product__name")
            .annotate(units_sold=Sum("quantity"), revenue=Sum("subtotal"))
            .order_by("-revenue")[:5]
        )
        top_products_list = [
            {
                "product_id": p["variant__product__id"],
                "name":       p["variant__product__name"],
                "units_sold": p["units_sold"],
                "revenue":    float(p["revenue"]),
            }
            for p in top_products
        ]

        # ── Métodos de pago ────────────────────────────────
        payment_methods = (
            qs.values("payment_method")
            .annotate(count=Count("id"), total=Sum("total"))
            .order_by("-count")
        )
        payment_list = [
            {"payment_method": p["payment_method"], "count": p["count"], "total": float(p["total"] or 0)}
            for p in payment_methods
        ]

        # ── Ventas por categoría ───────────────────────────
        by_category = (
            SaleItem.objects.filter(sale__in=qs)
            .values("variant__product__category__name")
            .annotate(revenue=Sum("subtotal"), count=Count("id"))
            .order_by("-revenue")
        )
        category_list = [
            {"category": c["variant__product__category__name"] or "Sin categoría",
             "revenue":  float(c["revenue"]),
             "count":    c["count"]}
            for c in by_category
        ]

        # ── Resumen de inventario (queries de agregación) ──
        inv_agg = ProductVariant.objects.filter(is_active=True).aggregate(
            total_stock=Sum("stock"),
            out_of_stock_count=Count("id", filter=Q(stock=0)),
        )
        total_stock = inv_agg["total_stock"] or 0
        out_of_stock_count = inv_agg["out_of_stock_count"] or 0
        low_stock_count = ProductVariant.objects.filter(
            is_active=True, stock__gt=0, stock__lte=F("product__min_stock")
        ).count()
        inventory_value = float(
            ProductVariant.objects.filter(is_active=True, stock__gt=0)
            .annotate(line_value=F("stock") * F("product__price"))
            .aggregate(t=Sum("line_value"))["t"] or 0
        )

        return Response({
            "total_revenue":       total_revenue,
            "total_sales":         total_sales,
            "daily_average":       daily_avg,
            "sales_by_day":        sales_by_day,
            "top_products":        top_products_list,
            "payment_methods":     payment_list,
            "sales_by_category":   category_list,
            # Inventario
            "total_stock":         total_stock,
            "out_of_stock":        out_of_stock_count,
            "low_stock":           low_stock_count,
            "inventory_value":     round(inventory_value, 2),
        })



class ExportSalesView(APIView):
    """Exportar ventas - devuelve base64 JSON para evitar corrupcion binaria en proxies."""
    permission_classes = [CanViewOnly]

    def get(self, request):
        import base64
        from datetime import date, timedelta

        qs = Sale.objects.select_related("sold_by", "customer").prefetch_related(
            "items__variant__product"
        ).order_by("-created_at")

        days      = request.query_params.get("days")
        from_date = request.query_params.get("from_date")
        to_date   = request.query_params.get("to_date")

        if days:
            try:
                cutoff = date.today() - timedelta(days=int(days))
                qs = qs.filter(created_at__date__gte=cutoff)
            except (ValueError, TypeError):
                pass
        else:
            if from_date:
                qs = qs.filter(created_at__date__gte=from_date)
            if to_date:
                qs = qs.filter(created_at__date__lte=to_date)

        fmt_val = request.query_params.get("file_format", "xlsx").lower()

        # CSV
        if fmt_val == "csv":
            import csv as csv_mod
            buffer = io.StringIO()
            writer = csv_mod.writer(buffer)
            writer.writerow(["Numero", "Fecha", "Cliente", "Subtotal", "Total",
                             "Metodo Pago", "Vendedor", "Puntos Usados", "Puntos Ganados"])
            for sale in qs:
                writer.writerow([
                    sale.number,
                    sale.created_at.strftime("%Y-%m-%d %H:%M"),
                    sale.customer.name if sale.customer else "",
                    float(sale.subtotal),
                    float(sale.total),
                    _payment_label(sale.payment_method),
                    sale.sold_by.get_full_name() if sale.sold_by else "",
                    sale.points_used,
                    sale.points_earned,
                ])
            b64 = base64.b64encode(buffer.getvalue().encode("utf-8")).decode("utf-8")
            return Response({"format": "csv", "filename": f"ventas-{date.today()}.csv", "data": b64})

        # Excel - base64 en JSON evita corrupcion binaria
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({"detail": "openpyxl no instalado. Usa format=csv."}, status=500)

        try:
            from apps.users.models import StoreConfig
            store_name = StoreConfig.get_config().name
        except Exception:
            store_name = "LUMA Store"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas"

        ws.merge_cells("A1:J1")
        hc = ws["A1"]
        hc.value     = f"Reporte de Ventas - {store_name}"
        hc.font      = Font(bold=True, size=14, color="FFFFFF")
        hc.alignment = Alignment(horizontal="center")
        hc.fill      = PatternFill("solid", fgColor="0D8585")

        ws.append([])
        cols = ["#", "Numero", "Fecha", "Cliente", "Subtotal", "Descuento",
                "Total", "Metodo Pago", "Vendedor", "Pts Ganados"]
        ws.append(cols)
        for cell in ws[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0D8585")

        for i, sale in enumerate(qs, 1):
            discount = float(sale.subtotal) - float(sale.total)
            ws.append([
                i,
                sale.number,
                sale.created_at.strftime("%Y-%m-%d %H:%M"),
                sale.customer.name if sale.customer else "-",
                float(sale.subtotal),
                round(discount, 2),
                float(sale.total),
                _payment_label(sale.payment_method),
                sale.sold_by.get_full_name() if sale.sold_by else "-",
                sale.points_earned,
            ])

        for col in ws.columns:
            real_cells = [c for c in col if c.__class__.__name__ != "MergedCell"]
            if not real_cells:
                continue
            max_len = max((len(str(c.value or "")) for c in real_cells), default=10)
            ws.column_dimensions[real_cells[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
        return Response({"format": "xlsx", "filename": f"ventas-{date.today()}.xlsx", "data": b64})


# ─────────────────────────────────────────────────────────────
#  HELPERS COMPARTIDOS
# ─────────────────────────────────────────────────────────────

def _get_store_name():
    try:
        from apps.users.models import StoreConfig
        return StoreConfig.get_config().name
    except Exception:
        return "LUMA Store"


def _xlsx_header(ws, title, cols, col_count):
    """Escribe encabezado con título y fila de columnas con estilo."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells(f"A1:{chr(64 + col_count)}1")
    hc = ws["A1"]
    hc.value = title
    hc.font = Font(bold=True, size=13, color="FFFFFF")
    hc.fill = PatternFill("solid", fgColor="0D8585")
    hc.alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(cols)
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D8585")


def _xlsx_autowidth(ws):
    for col in ws.columns:
        real_cells = [c for c in col if c.__class__.__name__ != "MergedCell"]
        if not real_cells:
            continue
        max_len = max((len(str(c.value or "")) for c in real_cells), default=10)
        ws.column_dimensions[real_cells[0].column_letter].width = min(max_len + 4, 40)


def _xlsx_to_b64(wb):
    import base64
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def _csv_to_b64(rows, headers):
    import base64, csv as csv_mod
    buf = io.StringIO()
    w = csv_mod.writer(buf)
    w.writerow(headers)
    w.writerows(rows)
    return base64.b64encode(buf.getvalue().encode("utf-8")).decode("utf-8")


# ─────────────────────────────────────────────────────────────
#  REPORTE DE INVENTARIO (ampliado con filtros)
# ─────────────────────────────────────────────────────────────

class InventoryReportView(APIView):
    """Reporte del estado del inventario con filtros opcionales."""
    permission_classes = [CanViewOnly]

    def get(self, request):
        category_id = request.query_params.get("category")
        status_filter = request.query_params.get("status")  # active | inactive | out

        qs = Product.objects.prefetch_related("variants").select_related("category")
        if category_id:
            qs = qs.filter(category_id=category_id)
        if status_filter:
            qs = qs.filter(status=status_filter)
        else:
            qs = qs.filter(status__in=["active", "out"])

        products_list = []
        total_cost_value = 0
        total_sale_value = 0
        total_stock = 0
        out_of_stock_count = 0
        low_stock_count = 0

        for p in qs:
            stock = sum(v.stock for v in p.variants.all())
            cost_val = float(p.cost or 0) * stock
            sale_val = float(p.price or 0) * stock
            total_cost_value += cost_val
            total_sale_value += sale_val
            total_stock += stock
            if stock == 0:
                out_of_stock_count += 1
            elif stock <= p.min_stock:
                low_stock_count += 1
            products_list.append({
                "id": p.id,
                "name": p.name,
                "category": p.category.name if p.category else "Sin categoría",
                "status": p.status,
                "stock": stock,
                "price": float(p.price or 0),
                "cost": float(p.cost or 0),
                "sale_value": round(sale_val, 2),
                "cost_value": round(cost_val, 2),
                "min_stock": p.min_stock,
            })

        # Categorías disponibles (solo activas, para los selects del frontend)
        from apps.inventory.models import Category
        categories = list(Category.objects.filter(is_active=True).values("id", "name").order_by("name"))

        return Response({
            "total_products": len(products_list),
            "total_stock": total_stock,
            "out_of_stock": out_of_stock_count,
            "low_stock": low_stock_count,
            "total_cost_value": round(total_cost_value, 2),
            "total_sale_value": round(total_sale_value, 2),
            "potential_margin": round(total_sale_value - total_cost_value, 2),
            "products": products_list,
            "categories": categories,
        })


# ─────────────────────────────────────────────────────────────
#  REPORTE DE PRODUCTOS
# ─────────────────────────────────────────────────────────────

class ProductReportView(APIView):
    """Top productos, por categoría y productos lentos."""
    permission_classes = [CanViewOnly]

    def get(self, request):
        from datetime import timedelta
        from django.db.models import Q

        days = int(request.query_params.get("days", 30))
        category_id = request.query_params.get("category")
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days - 1)

        sale_items_qs = SaleItem.objects.filter(
            sale__created_at__date__gte=start_date,
            sale__created_at__date__lte=end_date,
        )
        if category_id:
            sale_items_qs = sale_items_qs.filter(variant__product__category_id=category_id)

        # Top productos por ingresos
        top_revenue = (
            sale_items_qs
            .values("variant__product__id", "variant__product__name", "variant__product__category__name")
            .annotate(units_sold=Sum("quantity"), revenue=Sum("subtotal"))
            .order_by("-revenue")[:10]
        )
        top_revenue_list = [
            {
                "product_id": p["variant__product__id"],
                "name": p["variant__product__name"],
                "category": p["variant__product__category__name"] or "Sin categoría",
                "units_sold": p["units_sold"],
                "revenue": float(p["revenue"]),
            }
            for p in top_revenue
        ]

        # Top productos por unidades
        top_units = (
            sale_items_qs
            .values("variant__product__id", "variant__product__name")
            .annotate(units_sold=Sum("quantity"), revenue=Sum("subtotal"))
            .order_by("-units_sold")[:10]
        )
        top_units_list = [
            {
                "product_id": p["variant__product__id"],
                "name": p["variant__product__name"],
                "units_sold": p["units_sold"],
                "revenue": float(p["revenue"]),
            }
            for p in top_units
        ]

        # Ventas por categoría
        by_category = (
            sale_items_qs
            .values("variant__product__category__name")
            .annotate(revenue=Sum("subtotal"), units_sold=Sum("quantity"))
            .order_by("-revenue")
        )
        category_list = [
            {
                "category": c["variant__product__category__name"] or "Sin categoría",
                "revenue": float(c["revenue"]),
                "units_sold": c["units_sold"],
            }
            for c in by_category
        ]

        # Productos sin movimiento (lentos) — en stock pero sin ventas en el período
        sold_product_ids = set(
            sale_items_qs.values_list("variant__product_id", flat=True).distinct()
        )
        slow_qs = Product.objects.prefetch_related("variants").filter(
            status__in=["active", "out"]
        )
        if category_id:
            slow_qs = slow_qs.filter(category_id=category_id)
        slow_movers = []
        for p in slow_qs:
            stock = sum(v.stock for v in p.variants.all())
            if p.id not in sold_product_ids and stock > 0:
                slow_movers.append({
                    "product_id": p.id,
                    "name": p.name,
                    "stock": stock,
                    "price": float(p.price or 0),
                })
        slow_movers = slow_movers[:10]

        from apps.inventory.models import Category
        categories = list(Category.objects.filter(is_active=True).values("id", "name").order_by("name"))

        return Response({
            "top_by_revenue": top_revenue_list,
            "top_by_units": top_units_list,
            "by_category": category_list,
            "slow_movers": slow_movers,
            "categories": categories,
            "days": days,
        })


# ─────────────────────────────────────────────────────────────
#  REPORTE DE CLIENTES
# ─────────────────────────────────────────────────────────────

class CustomerReportView(APIView):
    """Segmentación de clientes, top compradores y KPIs."""
    permission_classes = [CanViewOnly]

    def get(self, request):
        from datetime import timedelta
        from apps.customers.models import Customer

        days = int(request.query_params.get("days", 30))
        segment_filter = request.query_params.get("segment")

        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days - 1)

        now = timezone.now()
        customers_qs = Customer.objects.annotate(
            sales_last_15=Count(
                "sales", filter=Q(sales__created_at__gte=now - timedelta(days=15))
            ),
            sales_last_60=Count(
                "sales", filter=Q(sales__created_at__gte=now - timedelta(days=60))
            ),
        )

        # KPIs generales
        total_customers = customers_qs.count()
        new_in_period = customers_qs.filter(created_at__date__gte=start_date).count()

        # Segmentos calculados a partir de anotaciones (sin queries por cliente)
        segments = {"new": 0, "frequent": 0, "regular": 0, "inactive": 0}
        for c in customers_qs.values("created_at", "sales_last_15", "sales_last_60"):
            last_15 = c["sales_last_15"] or 0
            last_60 = c["sales_last_60"] or 0
            is_new = c["created_at"] >= now - timedelta(days=15)
            if last_15 == 1 and is_new:
                segments["new"] += 1
            elif last_60 > 3:
                segments["frequent"] += 1
            elif last_60 >= 1:
                segments["regular"] += 1
            else:
                segments["inactive"] += 1
        segments_list = [
            {"segment": k, "count": v}
            for k, v in sorted(segments.items(), key=lambda x: -x[1])
        ]

        # Top clientes por gasto en el período
        top_buyers = (
            Sale.objects
            .filter(customer__isnull=False, created_at__date__gte=start_date)
            .values("customer__id", "customer__name", "customer__phone")
            .annotate(total_spent=Sum("total"), purchase_count=Count("id"))
            .order_by("-total_spent")[:10]
        )
        top_buyers_list = [
            {
                "customer_id": b["customer__id"],
                "name": b["customer__name"],
                "phone": b["customer__phone"] or "",
                "total_spent": float(b["total_spent"]),
                "purchase_count": b["purchase_count"],
            }
            for b in top_buyers
        ]

        # Puntos totales activos
        total_points = customers_qs.aggregate(t=Sum("points"))["t"] or 0

        # Ingresos del período generados por clientes identificados
        period_revenue_identified = float(
            Sale.objects.filter(
                customer__isnull=False, created_at__date__gte=start_date
            ).aggregate(t=Sum("total"))["t"] or 0
        )
        period_revenue_total = float(
            Sale.objects.filter(
                created_at__date__gte=start_date
            ).aggregate(t=Sum("total"))["t"] or 0
        )

        # Tasa de retención simple: clientes con más de 1 compra
        repeat_customers = customers_qs.annotate(
            cnt=Count("sales", distinct=True)
        ).filter(cnt__gt=1).count()
        retention_rate = round(repeat_customers / total_customers * 100, 1) if total_customers else 0

        # Clientes que compraron en el período
        active_in_period = Sale.objects.filter(
            customer__isnull=False, created_at__date__gte=start_date
        ).values("customer").distinct().count()

        return Response({
            "total_customers": total_customers,
            "new_in_period": new_in_period,
            "active_in_period": active_in_period,
            "total_points": total_points,
            "retention_rate": retention_rate,
            "period_revenue_identified": period_revenue_identified,
            "period_revenue_total": period_revenue_total,
            "segments": segments_list,
            "top_buyers": top_buyers_list,
            "days": days,
        })


# ─────────────────────────────────────────────────────────────
#  REPORTE DE CAJA
# ─────────────────────────────────────────────────────────────

class CashReportView(APIView):
    """Resumen de sesiones de caja, por método de pago y movimientos."""
    permission_classes = [CanViewOnly]

    def get(self, request):
        from datetime import timedelta
        from apps.cash.models import CashSession, CashMovement
        from django.db.models import Avg

        days = int(request.query_params.get("days", 30))
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days - 1)

        # Anota ingresos/egresos/devoluciones en una sola query con Case/When
        sessions = (
            CashSession.objects
            .filter(date__gte=start_date, date__lte=end_date)
            .annotate(
                income=Sum(
                    Case(When(movements__type="income", then="movements__amount"),
                         default=Value(0), output_field=DecimalField())
                ),
                expense=Sum(
                    Case(When(movements__type="expense", then="movements__amount"),
                         default=Value(0), output_field=DecimalField())
                ),
                refund=Sum(
                    Case(When(movements__type="refund", then="movements__amount"),
                         default=Value(0), output_field=DecimalField())
                ),
            )
            .order_by("-date")
        )

        sessions_data = []
        total_income = total_expense = total_refund = 0.0
        for s in sessions:
            inc = float(s.income or 0)
            exp = float(s.expense or 0)
            ref = float(s.refund or 0)
            total_income  += inc
            total_expense += exp
            total_refund  += ref
            sessions_data.append({
                "id": s.id,
                "date": str(s.date),
                "status": s.status,
                "opening_amount": float(s.opening_amount),
                "closing_amount": float(s.closing_amount) if s.closing_amount else None,
                "income": inc,
                "expense": exp,
                "refund": ref,
                "net": inc - exp - ref,
            })

        # Por método de pago
        movements_qs = CashMovement.objects.filter(
            session__date__gte=start_date, session__date__lte=end_date, type="income"
        )
        by_payment = (
            movements_qs
            .values("payment_method")
            .annotate(total=Sum("amount"), count=Count("id"))
            .order_by("-total")
        )
        payment_list = [
            {
                "payment_method": p["payment_method"],
                "total": float(p["total"]),
                "count": p["count"],
            }
            for p in by_payment
        ]

        return Response({
            "total_sessions": len(sessions_data),
            "total_income": round(total_income, 2),
            "total_expense": round(total_expense, 2),
            "total_refund": round(total_refund, 2),
            "net_cash": round(total_income - total_expense - total_refund, 2),
            "sessions": sessions_data,
            "by_payment_method": payment_list,
            "days": days,
        })


# ─────────────────────────────────────────────────────────────
#  EXPORTACIONES
# ─────────────────────────────────────────────────────────────

class ExportInventoryView(APIView):
    permission_classes = [CanViewOnly]

    def get(self, request):
        from datetime import date as date_cls
        category_id = request.query_params.get("category")
        status_filter = request.query_params.get("status")
        fmt_val = request.query_params.get("file_format", "xlsx").lower()

        qs = Product.objects.prefetch_related("variants").select_related("category")
        if category_id:
            qs = qs.filter(category_id=category_id)
        if status_filter:
            qs = qs.filter(status=status_filter)
        else:
            qs = qs.filter(status__in=["active", "out"])

        rows = []
        for i, p in enumerate(qs, 1):
            stock = sum(v.stock for v in p.variants.all())
            rows.append([
                i, p.name,
                p.category.name if p.category else "Sin categoría",
                p.status, stock,
                float(p.price or 0), float(p.cost or 0),
                round(float(p.price or 0) * stock, 2),
                p.min_stock,
            ])

        headers = ["#", "Producto", "Categoría", "Estado", "Stock",
                   "Precio Venta", "Costo", "Valor Stock", "Stock Mínimo"]
        filename = f"inventario-{date_cls.today()}"

        if fmt_val == "csv":
            b64 = _csv_to_b64(rows, headers)
            return Response({"format": "csv", "filename": f"{filename}.csv", "data": b64})

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return Response({"detail": "openpyxl no instalado."}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"
        _xlsx_header(ws, f"Inventario - {_get_store_name()}", headers, len(headers))
        for row in rows:
            ws.append(row)
        _xlsx_autowidth(ws)
        return Response({"format": "xlsx", "filename": f"{filename}.xlsx", "data": _xlsx_to_b64(wb)})


class ExportProductsView(APIView):
    permission_classes = [CanViewOnly]

    def get(self, request):
        from datetime import date as date_cls, timedelta

        days = int(request.query_params.get("days", 30))
        category_id = request.query_params.get("category")
        fmt_val = request.query_params.get("file_format", "xlsx").lower()
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days - 1)

        sale_items_qs = SaleItem.objects.filter(
            sale__created_at__date__gte=start_date,
            sale__created_at__date__lte=end_date,
        )
        if category_id:
            sale_items_qs = sale_items_qs.filter(variant__product__category_id=category_id)

        top = (
            sale_items_qs
            .values("variant__product__id", "variant__product__name", "variant__product__category__name")
            .annotate(units_sold=Sum("quantity"), revenue=Sum("subtotal"))
            .order_by("-revenue")
        )
        rows = [
            [i, p["variant__product__name"],
             p["variant__product__category__name"] or "Sin categoría",
             p["units_sold"], float(p["revenue"])]
            for i, p in enumerate(top, 1)
        ]
        headers = ["#", "Producto", "Categoría", "Unidades Vendidas", "Ingresos"]
        filename = f"productos-{date_cls.today()}"

        if fmt_val == "csv":
            b64 = _csv_to_b64(rows, headers)
            return Response({"format": "csv", "filename": f"{filename}.csv", "data": b64})

        try:
            import openpyxl
        except ImportError:
            return Response({"detail": "openpyxl no instalado."}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        _xlsx_header(ws, f"Reporte Productos - {_get_store_name()} (últimos {days} días)", headers, len(headers))
        for row in rows:
            ws.append(row)
        _xlsx_autowidth(ws)
        return Response({"format": "xlsx", "filename": f"{filename}.xlsx", "data": _xlsx_to_b64(wb)})


class ExportCustomersView(APIView):
    permission_classes = [CanViewOnly]

    def get(self, request):
        from datetime import date as date_cls, timedelta
        from apps.customers.models import Customer

        days = int(request.query_params.get("days", 30))
        fmt_val = request.query_params.get("file_format", "xlsx").lower()
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days - 1)

        # Pre-agrega ventas del período en una sola query
        period_map = {
            row["customer_id"]: row
            for row in Sale.objects.filter(
                customer__isnull=False,
                created_at__date__gte=start_date,
            )
            .values("customer_id")
            .annotate(
                period_count=Count("id"),
                period_total=Sum("total"),
            )
        }

        customers_qs = Customer.objects.annotate(
            total_spent=Sum("sales__total"),
            num_sales=Count("sales", distinct=True),
            sales_last_15=Count(
                "sales",
                filter=Q(sales__created_at__gte=timezone.now() - timedelta(days=15)),
            ),
            sales_last_60=Count(
                "sales",
                filter=Q(sales__created_at__gte=timezone.now() - timedelta(days=60)),
            ),
        ).order_by("name")

        rows = []
        now = timezone.now()
        for c in customers_qs:
            # Segmento calculado desde anotaciones — 0 queries extra
            last_15 = c.sales_last_15 or 0
            last_60 = c.sales_last_60 or 0
            is_new = c.created_at >= now - timedelta(days=15)
            if last_15 == 1 and is_new:
                seg = "Nuevo"
            elif last_60 > 3:
                seg = "Frecuente"
            elif last_60 >= 1:
                seg = "Regular"
            else:
                seg = "Inactivo"
            p = period_map.get(c.id, {})
            rows.append([
                c.name, c.phone or "", c.email or "",
                seg, c.num_sales or 0,
                float(c.total_spent or 0),
                c.points or 0,
                p.get("period_count", 0),
                float(p.get("period_total") or 0),
            ])

        headers = ["Nombre", "Teléfono", "Email", "Segmento",
                   "Total Compras", "Total Gastado", "Puntos",
                   f"Compras ({days}d)", f"Gasto ({days}d)"]
        filename = f"clientes-{date_cls.today()}"

        if fmt_val == "csv":
            b64 = _csv_to_b64(rows, headers)
            return Response({"format": "csv", "filename": f"{filename}.csv", "data": b64})

        try:
            import openpyxl
        except ImportError:
            return Response({"detail": "openpyxl no instalado."}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"
        _xlsx_header(ws, f"Reporte Clientes - {_get_store_name()}", headers, len(headers))
        for row in rows:
            ws.append(row)
        _xlsx_autowidth(ws)
        return Response({"format": "xlsx", "filename": f"{filename}.xlsx", "data": _xlsx_to_b64(wb)})


class ExportCashView(APIView):
    permission_classes = [CanViewOnly]

    def get(self, request):
        from datetime import date as date_cls, timedelta
        from apps.cash.models import CashSession

        days = int(request.query_params.get("days", 30))
        fmt_val = request.query_params.get("file_format", "xlsx").lower()
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days - 1)

        # Una sola query con Case/When — mismo patrón que CashReportView
        sessions = (
            CashSession.objects
            .filter(date__gte=start_date, date__lte=end_date)
            .annotate(
                income=Sum(
                    Case(When(movements__type="income", then="movements__amount"),
                         default=Value(0), output_field=DecimalField())
                ),
                expense=Sum(
                    Case(When(movements__type="expense", then="movements__amount"),
                         default=Value(0), output_field=DecimalField())
                ),
                refund=Sum(
                    Case(When(movements__type="refund", then="movements__amount"),
                         default=Value(0), output_field=DecimalField())
                ),
            )
            .order_by("-date")
        )

        rows = []
        for i, s in enumerate(sessions, 1):
            inc = float(s.income or 0)
            exp = float(s.expense or 0)
            ref = float(s.refund or 0)
            rows.append([
                i, str(s.date), s.status,
                float(s.opening_amount),
                float(s.closing_amount) if s.closing_amount else "",
                inc, exp, ref, round(inc - exp - ref, 2),
            ])

        headers = ["#", "Fecha", "Estado", "Apertura", "Cierre",
                   "Ingresos", "Gastos", "Devoluciones", "Neto"]
        filename = f"caja-{date_cls.today()}"

        if fmt_val == "csv":
            b64 = _csv_to_b64(rows, headers)
            return Response({"format": "csv", "filename": f"{filename}.csv", "data": b64})

        try:
            import openpyxl
        except ImportError:
            return Response({"detail": "openpyxl no instalado."}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Caja"
        _xlsx_header(ws, f"Reporte Caja - {_get_store_name()}", headers, len(headers))
        for row in rows:
            ws.append(row)
        _xlsx_autowidth(ws)
        return Response({"format": "xlsx", "filename": f"{filename}.xlsx", "data": _xlsx_to_b64(wb)})
